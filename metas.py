from openpyxl import load_workbook as lw
from time import strftime as st
from tkinter import messagebox
from functools import cache

dia = st('%Y-%m-%d 00:00:00')
mes = st('%Y%m')

  
class Metas():
    def __init__(self):
        self.soma = float()
        self.basket = float()
        self.blz = float()
        self.cba = float()
        self.tec = float()
        self.fem = float()
        self.inf = float()
        self.mas = float()
        self.mdc = float()
        self.biju = float()
        self.rel = float()
        try:
            self.planilha = lw(f"metas atualizadas//metas{mes}.xlsx")
            self.ws = self.planilha.active
            self.grupo()
        except:
            self.AA()

    def grupo(self):
        for i in range(4, 5000):
            diaP = self.ws[f'A{i}'].value
            valor = self.ws[f'D{i}'].value
            fl = self.ws[f'B{i}'].value
            grupo = self.ws[f"C{i}"].value
            diaP = str(diaP)
            if diaP == dia:
                if fl == 'L062':
                    if grupo == 'Basket':
                        self.basket = valor
                        self.soma += valor
                    if grupo == 'Beleza':
                        self.blz = valor
                        self.soma += valor
                    if grupo == 'Calçados':
                        self.cba = valor
                        self.soma += valor
                    if grupo == 'Eletrônicos':
                        self.tec = valor
                        self.soma += valor
                    if grupo == 'Feminino':
                        self.fem = valor
                        self.soma += valor
                    if grupo == 'Infantil':
                        self.inf = valor
                        self.soma += valor
                    if grupo == 'Masculino':
                        self.masc = valor
                        self.soma += valor
                    if grupo == 'Moda Casa':
                        self.mdc = valor
                        self.soma += valor
                    if grupo == 'Óculos e Bijuteria':
                        self.biju = valor
                        self.soma += valor
                    if grupo == 'Relógios':
                        self.rel = valor
                        self.soma += valor
                        
    def AA(self):
        ano = int(st('%Y'))
        ly = st(f'{ano}%m')
        ca = st('%m-%d')
        plnAA = lw(f'metas//metas{ly}.xlsx')
        wa = plnAA.active

        for i in range(4, 600):
            diaA = wa[f'B{i}'].value
            valor = wa[f'C{i}'].value
            dco = wa[f'D{i}'].value
            if diaA != None:
                if ca in diaA:
                    if dco == 'Basket':
                        self.basket = valor
                        self.soma += valor
                    if dco == 'Beleza':
                        self.blz = valor
                        self.soma += valor
                    if dco == 'Calçados':
                        self.cba = valor
                        self.soma += valor
                    if dco == 'Eletrônicos':
                        self.tec = valor
                        self.soma += valor
                    if dco == 'Feminino':
                        self.fem = valor
                        self.soma += valor
                    if dco == 'Infantil':
                        self.inf = valor
                        self.soma += valor
                    if dco == 'Masculino':
                        self.masc = valor
                        self.soma += valor
                    if dco == 'Moda Casa':
                        self.mdc = valor
                        self.soma += valor
                    if dco == 'Óculos e Bijuteria':
                        self.biju = valor
                        self.soma += valor
                    if dco == 'Relógios':
                        self.rel = valor
                        self.soma += valor
Metas()
