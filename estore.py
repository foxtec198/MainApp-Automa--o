from openpyxl import load_workbook as lw
from time import strftime as st, sleep

dia = int(st("%d"))
data = st("%d/%m/%Y")
pln = lw(f"metas.xlsx")
ws = pln.active 
soma = 0
total = 0

def L(lj):
    global soma
    soma = 0
    linhas = 6000
    for i in range(4, linhas):
        diaP = ws[f"A{i}"].value
        loja = ws[f"B{i}"].value
        dco = ws[f"C{i}"].value
        valor = ws[f"D{i}"].value
        # Retira celulas vazias 
        if diaP != None:
            if diaP == dia:
                if loja == f'L{lj}':
                    soma += valor
    soma = soma * 0.01
    soma = round(soma)
    
L("025")
l25 = soma

L("051")
l51 = soma

L("054")
l54 = soma

L("056")
l56 = soma

L("057")
l57 = soma

L("059")
l59 = soma

L("061")
l61 = soma

L("062")
l62 = soma

L("194")
l194 = soma

L("201")
l201 = soma

L("287")
l287 = soma

L("306")
l306 = soma

L("313")
l313 = soma

L("316")
l316 = soma

L("326")
l326 = soma

L("391")
l391 = soma

L("393")
l393 = soma

total = l25 + l51 + l54 + l56 + l57 +l59 +l61 + l62 + l194 + l201 + l287 + l306+l313 + l316 + l326 + l391 + l393

text = f'''
*METAS DIÁRIAS E-STORE CTB*
 🗓️ {data}

_*Meta//Realizado*_

L025: R${l25}//
L051: R${l51}//
L054: R${l54}//
L056: R${l56}//
L057: R${l57}//
L059: R${l59}//
L061: R${l61}//
L062: R${l62}//
L194: R${l194}//
L201: R${l201}//
L287: R${l287}//
L306: R${l306}//
L313: R${l313}//
L316: R${l316}//
L326: R${l326}//
L391: R${l391}//
L393: R${l393}//


*TOTAL: R$ {total}*
*REALIZADO: R$ 0*


*FEZ A VENDA ATUALIZA!* 😝
'''
print(text)
sleep(30)
